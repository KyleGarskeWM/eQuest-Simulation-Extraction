#!/usr/bin/env python3
"""Extract BEPS, LV-B, LV-D, LV-I, LS-A, LV-M, ES-D, and PS-H report data from an eQuest .SIM file."""
from __future__ import annotations
import argparse
import io
import json
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List
try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency for safer workbook writes
    load_workbook = None
END_USE_COLUMNS = [
    "LIGHTS",
    "TASK LIGHTS",
    "MISC EQUIP",
    "SPACE HEATING",
    "SPACE COOLING",
    "HEAT REJECT",
    "PUMPS & AUX",
    "VENT FANS",
    "REFRIG DISPLAY",
    "HT PUMP SUPPLEM",
    "DOMEST HOT WTR",
    "EXT USAGE",
    "TOTAL",
]
LV_D_COLUMNS = [
    "orientation",
    "avg_u_value_windows",
    "avg_u_value_walls",
    "avg_u_value_walls_plus_windows",
    "window_area",
    "wall_area",
    "window_plus_wall_area",
]
LV_D_UNITS = {
    "avg_u_value_windows": "BTU/HR-SQFT-F",
    "avg_u_value_walls": "BTU/HR-SQFT-F",
    "avg_u_value_walls_plus_windows": "BTU/HR-SQFT-F",
    "window_area": "SQFT",
    "wall_area": "SQFT",
    "window_plus_wall_area": "SQFT",
}
LV_D_TARGET_ORIENTATIONS = {
    "NORTH",
    "NORTH-EAST",
    "EAST",
    "SOUTH-EAST",
    "SOUTH",
    "SOUTH-WEST",
    "WEST",
    "NORTH-WEST",
    "FLOOR",
    "ROOF",
    "ALL WALLS",
    "WALLS+ROOFS",
    "UNDERGRND",
    "BUILDING",
}
NUMBER_PATTERN = re.compile(r"-?[\d,]+(?:\.\d*)?")
CONDITIONED_FLOOR_AREA_PATTERN = re.compile(r"CONDITIONED FLOOR AREA\s*=\s*([\d,]+(?:\.\d+)?)\s+SQFT", re.IGNORECASE)
LV_D_SUMMARY_ROW_PATTERN = re.compile(
    r"^\s*([A-Z][A-Z\-\+\s]+?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+"
    r"(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s*$"
)
LV_I_ROW_PATTERN = re.compile(
    r"^\s*(.+?)\s+(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s+(\d+)\s+(DELAYED|QUICK)\s+(\d+)\s*$"
)
LV_I_UVALUE_UNIT_PATTERN = re.compile(r"U-VALUE\s*\(([^\)]+)\)", re.IGNORECASE)
LS_A_LOAD_UNIT_PATTERN = re.compile(r"COOLING LOAD\s*\(([^\)]+)\)", re.IGNORECASE)
REPORT_HEADER_PATTERN = re.compile(r"REPORT-\s*([A-Z0-9\-]+)", re.IGNORECASE)
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS = {"m": MAIN_NS}
MASTER_ROOM_LIST_SHEET_XML_PATH = "xl/worksheets/sheet1.xml"
MASTER_ROOM_LIST_SPACE_START_ROW = 16
MASTER_ROOM_LIST_SPACE_MAX_ROWS = 298
UTILITY_RATES_SHEET_XML_PATH = "xl/worksheets/sheet7.xml"
ECM_DATA_SHEET_XML_PATH = "xl/worksheets/sheet11.xml"
ECM_DATA_MODEL_START_ROWS = {
    "BASELINE": 6,
    "PROPOSED": 17,
    "ECM-1": 28,
    "ECM-2": 39,
    "ECM-3": 50,
    "ECM-4": 61,
    "ECM-5": 72,
    "ECM-6": 83,
    "ECM-7": 94,
}
BEPS_TO_ECM_END_USE_COLUMNS = {
    "LIGHTS": "B",  # Internal Lighting
    "EXT USAGE": "C",  # External Lighting
    "SPACE HEATING": "D",
    "SPACE COOLING": "E",
    "PUMPS & AUX": "F",
    "VENT FANS": "H",  # Fans Interior
    "DOMEST HOT WTR": "J",
    "MISC EQUIP": "K",  # Receptacle Equipment
    "TASK LIGHTS": "L",  # Interior Lighting Process
    "REFRIG DISPLAY": "M",
    "HEAT REJECT": "Q",
}
ECM_OPTIONAL_BLANK_COLUMNS = ["G", "I", "N", "O", "P", "R", "S", "T"]
KBTU_PER_UNIT = {
    "KWH": 3.412,
    "THERM": 100.0,
    "KBTU": 1.0,
    "MBTU": 1000.0,
    "MMBTU": 1000.0,
    "BTU": 0.001,
}
def _clean_number(value: str) -> float:
    return float(value.replace(",", ""))
def _parse_values_line(line: str) -> tuple[str, List[float]]:
    stripped = line.strip()
    if not stripped:
        raise ValueError("Expected values line but got blank line.")
    unit = stripped.split()[0]
    numbers = [_clean_number(token) for token in NUMBER_PATTERN.findall(line)]
    if len(numbers) != len(END_USE_COLUMNS):
        raise ValueError(
            f"Expected {len(END_USE_COLUMNS)} numeric BEPS columns but found {len(numbers)} in line: {line!r}"
        )
    return unit, numbers


def detect_available_reports(sim_text: str) -> Dict[str, object]:
    """Return discovered REPORT-* identifiers from a SIM file."""
    discovered = []
    seen = set()
    for line in sim_text.splitlines():
        match = REPORT_HEADER_PATTERN.search(line)
        if not match:
            continue
        report_id = match.group(1).upper()
        if report_id in seen:
            continue
        seen.add(report_id)
        discovered.append(report_id)
    return {
        "available_reports": discovered,
        "has_beps": "BEPS" in seen,
        "has_lv_b": "LV-B" in seen,
        "has_es_d": "ES-D" in seen,
    }
def extract_beps_report(sim_text: str) -> Dict[str, object]:
    """Parse BEPS and return electricity/natural-gas totals for each end-use column."""
    lines = sim_text.splitlines()
    report_start = next((i for i, line in enumerate(lines) if "REPORT- BEPS" in line.upper()), None)
    if report_start is None:
        raise ValueError(
            "Could not find 'REPORT- BEPS' in the SIM file. "
            "The SIM may be truncated/cut off before BEPS; run with '--list-reports' to verify report availability."
        )
    section_lines = lines[report_start : report_start + 300]
    rows: Dict[str, Dict[str, object]] = {}
    idx = 0
    while idx < len(section_lines):
        line = section_lines[idx].strip()
        upper = line.upper()
        if upper.startswith("REPORT-") and idx > 0:
            break
        is_electric = "ELECTRICITY" in upper
        is_gas = "NATURAL-GAS" in upper or "NATURAL GAS" in upper
        if not (is_electric or is_gas):
            idx += 1
            continue
        row_name = " ".join(line.split())
        j = idx + 1
        while j < len(section_lines) and not section_lines[j].strip():
            j += 1
        if j >= len(section_lines):
            raise ValueError(f"Missing values line for BEPS row '{row_name}'.")
        unit, values = _parse_values_line(section_lines[j])
        rows[row_name] = {
            "fuel_type": "electricity" if is_electric else "natural_gas",
            "unit": unit,
            "values": dict(zip(END_USE_COLUMNS, values)),
        }
        idx = j + 1
    if not rows:
        raise ValueError("No electricity or natural-gas rows were parsed from BEPS.")
    totals = {
        "electricity": {col: 0.0 for col in END_USE_COLUMNS},
        "natural_gas": {col: 0.0 for col in END_USE_COLUMNS},
    }
    units = {"electricity": None, "natural_gas": None}
    for row in rows.values():
        fuel_type = row["fuel_type"]
        row_unit = row["unit"]
        if units[fuel_type] is None:
            units[fuel_type] = row_unit
        elif units[fuel_type] != row_unit:
            raise ValueError(
                f"Inconsistent units for {fuel_type}: saw both '{units[fuel_type]}' and '{row_unit}'."
            )
        for col, value in row["values"].items():
            totals[fuel_type][col] += value
    return {
        "report": "BEPS",
        "columns": END_USE_COLUMNS,
        "rows": rows,
        "totals_by_fuel": {
            "electricity": {"unit": units["electricity"], "by_end_use": totals["electricity"]},
            "natural_gas": {"unit": units["natural_gas"], "by_end_use": totals["natural_gas"]},
        },
    }
def extract_lv_b_spaces(sim_text: str) -> Dict[str, object]:
    """Extract unique LV-B spaces, grouping label, requested attributes, and conditioned floor area."""
    lines = sim_text.splitlines()
    in_lvb = False
    current_group = None
    spaces: Dict[str, Dict[str, object]] = {}
    for raw_line in lines:
        stripped = raw_line.strip()
        upper = stripped.upper()
        if "REPORT- LV-B" in upper:
            in_lvb = True
            continue
        if in_lvb and upper.startswith("REPORT-") and "REPORT- LV-B" not in upper:
            in_lvb = False
        if not in_lvb:
            continue
        if upper.startswith("SPACES ON FLOOR:"):
            current_group = stripped
            continue
        if (
            not stripped
            or upper.startswith("NUMBER OF SPACES")
            or "SPACE*FLOOR" in upper
            or upper.startswith("BUILDING TOTALS")
            or "SPACE" == upper
            or set(stripped) <= {"-", "=", "+"}
        ):
            continue
        parts = [part.strip() for part in re.split(r"\s{2,}", raw_line.strip()) if part.strip()]
        if len(parts) < 11:
            continue
        space_name = parts[0]
        _space_floor_multiplier = parts[1]
        space_type = parts[2]
        _azimuth = parts[3]
        lights = parts[4]
        people = parts[5]
        equip = parts[6]
        infiltration_method = parts[7]
        ach = parts[8]
        area_sqft = parts[9]
        volume_cuft = parts[10]
        if space_type not in {"INT", "EXT"}:
            continue
        try:
            float(_space_floor_multiplier)
            float(_azimuth)
            float(lights)
            float(people)
            float(equip)
            float(ach)
            float(area_sqft)
            float(volume_cuft)
        except ValueError:
            continue
        normalized_name = " ".join(space_name.split())
        if normalized_name in spaces:
            continue
        spaces[normalized_name] = {
            "group": current_group,
            "space_type": space_type,
            "lights_w_per_sqft": float(lights),
            "people": float(people),
            "equip_w_per_sqft": float(equip),
            "infiltration_method": infiltration_method,
            "ach": float(ach),
            "area_sqft": float(area_sqft),
            "volume_cuft": float(volume_cuft),
            "units": {
                "lights": "WATT/SQFT",
                "equip": "WATT/SQFT",
                "area": "SQFT",
                "volume": "CUFT",
                "ach": "ACH",
            },
        }
    if not spaces:
        raise ValueError("Could not parse any LV-B space rows from the SIM file.")
    conditioned_floor_area_match = CONDITIONED_FLOOR_AREA_PATTERN.search(sim_text)
    conditioned_floor_area = None
    if conditioned_floor_area_match:
        conditioned_floor_area = float(conditioned_floor_area_match.group(1).replace(",", ""))
    return {
        "report": "LV-B",
        "space_count": len(spaces),
        "conditioned_floor_area_sqft": conditioned_floor_area,
        "spaces": spaces,
    }
def extract_lv_d_report(sim_text: str) -> Dict[str, object]:
    """Extract only the final LV-D summary section by major orientation/category."""
    lines = sim_text.splitlines()
    in_lvd = False
    summary_rows: Dict[str, Dict[str, float]] = {}
    for raw_line in lines:
        stripped = raw_line.strip()
        upper = stripped.upper()
        if "REPORT- LV-D" in upper:
            in_lvd = True
            continue
        if in_lvd and upper.startswith("REPORT-") and "REPORT- LV-D" not in upper:
            in_lvd = False
        if not in_lvd:
            continue
        match = LV_D_SUMMARY_ROW_PATTERN.match(raw_line)
        if not match:
            continue
        (
            orientation,
            avg_u_value_windows,
            avg_u_value_walls,
            avg_u_value_walls_plus_windows,
            window_area,
            wall_area,
            window_plus_wall_area,
        ) = match.groups()
        normalized_orientation = " ".join(orientation.split())
        if normalized_orientation not in LV_D_TARGET_ORIENTATIONS:
            continue
        summary_rows[normalized_orientation] = {
            "avg_u_value_windows": float(avg_u_value_windows),
            "avg_u_value_walls": float(avg_u_value_walls),
            "avg_u_value_walls_plus_windows": float(avg_u_value_walls_plus_windows),
            "window_area": float(window_area),
            "wall_area": float(wall_area),
            "window_plus_wall_area": float(window_plus_wall_area),
        }
    if not summary_rows:
        raise ValueError("Could not parse LV-D summary rows from the SIM file.")
    missing = sorted(LV_D_TARGET_ORIENTATIONS.difference(summary_rows.keys()))
    return {
        "report": "LV-D",
        "columns": LV_D_COLUMNS,
        "units": LV_D_UNITS,
        "orientations": summary_rows,
        "missing_orientations": missing,
    }
def extract_lv_i_constructions(sim_text: str) -> Dict[str, object]:
    """Extract LV-I construction names with U-value and number of response factors."""
    lines = sim_text.splitlines()
    in_lvi = False
    constructions: Dict[str, Dict[str, object]] = {}
    section_lines: List[str] = []
    for raw_line in lines:
        stripped = raw_line.strip()
        upper = stripped.upper()
        if "REPORT- LV-I" in upper:
            in_lvi = True
            continue
        if in_lvi and upper.startswith("REPORT-") and "REPORT- LV-I" not in upper:
            in_lvi = False
        if not in_lvi:
            continue
        section_lines.append(raw_line)
        match = LV_I_ROW_PATTERN.match(raw_line)
        if not match:
            continue
        construction_name = " ".join(match.group(1).split())
        u_value = float(match.group(2))
        response_factors = int(match.group(6))
        constructions[construction_name] = {
            "u_value": u_value,
            "number_of_response_factors": response_factors,
        }
    if not constructions:
        raise ValueError("Could not parse any LV-I construction rows from the SIM file.")
    unit = "BTU/HR-SQFT-F"
    section_text = "\n".join(section_lines)
    unit_match = LV_I_UVALUE_UNIT_PATTERN.search(section_text)
    if unit_match:
        unit = unit_match.group(1).strip()
    return {
        "report": "LV-I",
        "u_value_unit": unit,
        "construction_count": len(constructions),
        "constructions": constructions,
    }
def extract_ls_a_peak_loads(sim_text: str, lv_b_result: Dict[str, object] | None = None) -> Dict[str, object]:
    """Extract LS-A cooling/heating peak loads and associate them with LV-B spaces."""
    if lv_b_result is None:
        lv_b_result = extract_lv_b_spaces(sim_text)
    lv_b_spaces = lv_b_result["spaces"]
    lines = sim_text.splitlines()
    in_lsa = False
    loads_by_space: Dict[str, Dict[str, float]] = {}
    units = "KBTU/HR"
    for raw_line in lines:
        stripped = raw_line.strip()
        upper = stripped.upper()
        if "REPORT- LS-A" in upper:
            in_lsa = True
            continue
        if in_lsa and upper.startswith("REPORT-") and "REPORT- LS-A" not in upper:
            in_lsa = False
        if not in_lsa:
            continue
        unit_match = LS_A_LOAD_UNIT_PATTERN.search(raw_line)
        if unit_match:
            units = unit_match.group(1).strip()
        if (
            not stripped
            or upper.startswith("SPACE NAME")
            or upper.startswith("MULTIPLIER")
            or set(stripped) <= {"-", "=", "+"}
        ):
            continue
        parts = [part.strip() for part in re.split(r"\s{2,}", stripped) if part.strip()]
        if len(parts) < 4:
            continue
        space_name = " ".join(parts[0].split())
        numeric_parts = []
        for token in parts[3:]:
            if re.fullmatch(r"-?\d+(?:\.\d+)?", token):
                numeric_parts.append(float(token))
        if len(numeric_parts) < 2:
            continue
        loads_by_space[space_name] = {
            "cooling_load": numeric_parts[0],
            "heating_load": numeric_parts[1],
        }
    if not loads_by_space:
        raise ValueError("Could not parse any LS-A space peak loads from the SIM file.")
    spaces_with_peak_loads: Dict[str, Dict[str, object]] = {}
    for space_name, space_data in lv_b_spaces.items():
        merged = dict(space_data)
        merged["peak_loads"] = {
            "cooling_load": loads_by_space.get(space_name, {}).get("cooling_load"),
            "heating_load": loads_by_space.get(space_name, {}).get("heating_load"),
            "units": units,
        }
        spaces_with_peak_loads[space_name] = merged
    return {
        "report": "LS-A",
        "load_units": units,
        "space_peak_loads": loads_by_space,
        "spaces_with_peak_loads": spaces_with_peak_loads,
    }
def extract_lv_m_conversions(sim_text: str) -> Dict[str, object]:
    """Extract conversion factors from LV-M and store them for future unit transforms."""
    lines = sim_text.splitlines()
    in_lvm = False
    conversions: Dict[str, Dict[str, float]] = {}
    for raw_line in lines:
        stripped = raw_line.strip()
        upper = stripped.upper()
        if "REPORT- LV-M" in upper:
            in_lvm = True
            continue
        if in_lvm and upper.startswith("REPORT-") and "REPORT- LV-M" not in upper:
            in_lvm = False
        if not in_lvm:
            continue
        parts = [part.strip() for part in re.split(r"\s{2,}", stripped) if part.strip()]
        # Expected: idx, english, factor1, metric, factor2, english2
        if len(parts) < 6:
            continue
        if not parts[0].isdigit():
            continue
        source_unit = parts[1]
        try:
            source_to_target = float(parts[2])
            target_to_source = float(parts[4])
        except ValueError:
            continue
        target_unit = parts[3]
        reverse_unit = parts[5]
        conversions.setdefault(source_unit, {})[target_unit] = source_to_target
        conversions.setdefault(target_unit, {})[reverse_unit] = target_to_source
    if not conversions:
        raise ValueError("Could not parse LV-M unit conversion rows from the SIM file.")
    return {
        "report": "LV-M",
        "conversions": conversions,
        "usage": "Use convert_value(value, from_unit, to_unit, conversions) for future transformations.",
    }
def convert_value(value: float, from_unit: str, to_unit: str, conversions: Dict[str, Dict[str, float]]) -> float:
    """Convert a value between units using LV-M conversion factors (supports chained conversions)."""
    if from_unit == to_unit:
        return value
    visited = set()
    queue: List[tuple[str, float]] = [(from_unit, value)]
    while queue:
        current_unit, current_value = queue.pop(0)
        if current_unit in visited:
            continue
        visited.add(current_unit)
        for next_unit, factor in conversions.get(current_unit, {}).items():
            next_value = current_value * factor
            if next_unit == to_unit:
                return next_value
            if next_unit not in visited:
                queue.append((next_unit, next_value))
    raise ValueError(f"No conversion path found from '{from_unit}' to '{to_unit}'.")


def _ensure_row(sheet_data: ET.Element, row_number: int) -> ET.Element:
    row = sheet_data.find(f"m:row[@r='{row_number}']", NS)
    if row is not None:
        return row
    row = ET.Element(f"{{{MAIN_NS}}}row", {"r": str(row_number)})
    inserted = False
    for existing in sheet_data.findall("m:row", NS):
        if int(existing.attrib["r"]) > row_number:
            sheet_data.insert(list(sheet_data).index(existing), row)
            inserted = True
            break
    if not inserted:
        sheet_data.append(row)
    return row


def _set_inline_string_cell(row: ET.Element, cell_ref: str, value: str, style: str | None = None) -> None:
    cell = row.find(f"m:c[@r='{cell_ref}']", NS)
    if cell is None:
        attrs = {"r": cell_ref}
        if style is not None:
            attrs["s"] = style
        cell = ET.SubElement(row, f"{{{MAIN_NS}}}c", attrs)
    elif style is not None and "s" not in cell.attrib:
        cell.attrib["s"] = style
    for child in list(cell):
        cell.remove(child)
    cell.attrib["t"] = "inlineStr"
    is_node = ET.SubElement(cell, f"{{{MAIN_NS}}}is")
    text_node = ET.SubElement(is_node, f"{{{MAIN_NS}}}t")
    text_node.text = value


def _set_numeric_cell(row: ET.Element, cell_ref: str, value: float | None, style: str | None = None) -> None:
    cell = row.find(f"m:c[@r='{cell_ref}']", NS)
    if cell is None:
        attrs = {"r": cell_ref}
        if style is not None:
            attrs["s"] = style
        cell = ET.SubElement(row, f"{{{MAIN_NS}}}c", attrs)
    elif style is not None and "s" not in cell.attrib:
        cell.attrib["s"] = style
    for child in list(cell):
        cell.remove(child)
    cell.attrib.pop("t", None)
    if value is not None:
        v_node = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
        v_node.text = f"{value:.6f}".rstrip("0").rstrip(".")


def _load_zip_file_map(workbook_path: Path) -> Dict[str, bytes]:
    with zipfile.ZipFile(workbook_path, "r") as workbook_zip:
        return {name: workbook_zip.read(name) for name in workbook_zip.namelist()}


def _save_zip_file_map(file_map: Dict[str, bytes], output_workbook_path: Path) -> None:
    with zipfile.ZipFile(output_workbook_path, "w", compression=zipfile.ZIP_DEFLATED) as dst_zip:
        for name, payload in file_map.items():
            dst_zip.writestr(name, payload)


def _parse_xml_with_registered_namespaces(xml_payload: bytes) -> ET.Element:
    for _, ns in ET.iterparse(io.BytesIO(xml_payload), events=("start-ns",)):
        prefix, uri = ns
        ET.register_namespace(prefix or "", uri)
    return ET.fromstring(xml_payload)


def _to_kbtu(value: float, from_unit: str) -> float:
    normalized_unit = from_unit.upper().strip()
    if normalized_unit not in KBTU_PER_UNIT:
        raise ValueError(f"Unsupported unit conversion to kBtu from '{from_unit}'.")
    return value * KBTU_PER_UNIT[normalized_unit]


def _excel_column_name(column_number: int) -> str:
    name = ""
    n = column_number
    while n > 0:
        n, rem = divmod(n - 1, 26)
        name = chr(65 + rem) + name
    return name


def _load_master_room_list_sheet(workbook_path: Path) -> ET.Element:
    with zipfile.ZipFile(workbook_path, "r") as workbook_zip:
        try:
            sheet_xml = workbook_zip.read(MASTER_ROOM_LIST_SHEET_XML_PATH)
        except KeyError as exc:
            raise ValueError("Could not find Master Room List worksheet XML in the workbook.") from exc
    return _parse_xml_with_registered_namespaces(sheet_xml)


def _read_cell_text(row: ET.Element, cell_ref: str) -> str:
    cell = row.find(f"m:c[@r='{cell_ref}']", NS)
    if cell is None:
        return ""
    if cell.attrib.get("t") == "inlineStr":
        text_parts = [node.text or "" for node in cell.findall(".//m:t", NS)]
        return "".join(text_parts)
    value_node = cell.find("m:v", NS)
    return value_node.text.strip() if value_node is not None and value_node.text else ""


def _read_cell_float(row: ET.Element, cell_ref: str) -> float | None:
    cell = row.find(f"m:c[@r='{cell_ref}']", NS)
    if cell is None:
        return None
    value_node = cell.find("m:v", NS)
    if value_node is None or value_node.text is None or not value_node.text.strip():
        return None
    try:
        return float(value_node.text.strip())
    except ValueError:
        return None


def _normalize_space_name(space_name: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", space_name.upper())


def extract_hourly_thermostat_setpoint_ranges(sim_text: str) -> Dict[str, object]:
    """Extract per-space min/max thermostat setpoint values from REPORT- HOURLY sections."""
    lines = sim_text.splitlines()
    in_hourly = False
    current_space_name = None
    setpoint_values_by_space: Dict[str, List[float]] = {}
    canonical_to_name: Dict[str, str] = {}
    for raw_line in lines:
        stripped = raw_line.strip()
        upper = stripped.upper()
        if "REPORT- HOURLY" in upper:
            in_hourly = True
            current_space_name = None
            continue
        if in_hourly and upper.startswith("REPORT-") and "REPORT- HOURLY" not in upper:
            in_hourly = False
            current_space_name = None
        if not in_hourly or not stripped:
            continue
        if upper.startswith("SPACE:"):
            current_space_name = stripped.split(":", 1)[1].strip()
            canonical = _normalize_space_name(current_space_name)
            canonical_to_name.setdefault(canonical, current_space_name)
            setpoint_values_by_space.setdefault(canonical, [])
            continue
        if current_space_name is None:
            continue
        if "THERMOSTAT SETPOINT" not in upper and not re.match(r"^\d+\s+", stripped):
            continue
        number_tokens = re.findall(r"-?\d+(?:\.\d+)?", stripped)
        if len(number_tokens) < 2:
            continue
        try:
            setpoint_value = float(number_tokens[1])
        except ValueError:
            continue
        canonical = _normalize_space_name(current_space_name)
        setpoint_values_by_space.setdefault(canonical, []).append(setpoint_value)
    spaces: Dict[str, Dict[str, float]] = {}
    for canonical_name, values in setpoint_values_by_space.items():
        if not values:
            continue
        display_name = canonical_to_name.get(canonical_name, canonical_name)
        spaces[display_name] = {
            "min_thermostat_setpoint_f": min(values),
            "max_thermostat_setpoint_f": max(values),
        }
    return {
        "report": "HOURLY",
        "space_count": len(spaces),
        "spaces": spaces,
    }


def _write_utility_rate_table_from_es_d(
    file_map: Dict[str, bytes],
    es_d_result: Dict[str, object],
) -> None:
    if UTILITY_RATES_SHEET_XML_PATH not in file_map:
        raise ValueError("Could not find Utilities worksheet XML in workbook.")
    utility_root = _parse_xml_with_registered_namespaces(file_map[UTILITY_RATES_SHEET_XML_PATH])
    utility_sheet_data = utility_root.find("m:sheetData", NS)
    if utility_sheet_data is None:
        raise ValueError("Utilities sheet is missing sheetData.")
    utility_rates = es_d_result["utility_rates"]
    elec_virtual_rate = None
    gas_virtual_rate = None
    for rate_data in utility_rates.values():
        unit = str(rate_data["unit"]).upper()
        if unit == "KWH" and elec_virtual_rate is None:
            elec_virtual_rate = float(rate_data["virtual_rate"])
        elif unit == "THERM" and gas_virtual_rate is None:
            gas_virtual_rate = float(rate_data["virtual_rate"])
    if elec_virtual_rate is None or gas_virtual_rate is None:
        raise ValueError("Could not find ES-D virtual rates for both electricity (KWH) and gas (THERM).")
    row_2 = _ensure_row(utility_sheet_data, 2)
    row_3 = _ensure_row(utility_sheet_data, 3)
    _set_inline_string_cell(row_2, "B2", "Electrical")
    _set_inline_string_cell(row_2, "C2", "kWh")
    _set_numeric_cell(row_2, "D2", elec_virtual_rate)
    _set_numeric_cell(row_2, "E2", elec_virtual_rate / KBTU_PER_UNIT["KWH"])
    _set_inline_string_cell(row_3, "B3", "Natural Gas")
    _set_inline_string_cell(row_3, "C3", "Therms")
    _set_numeric_cell(row_3, "D3", gas_virtual_rate)
    _set_numeric_cell(row_3, "E3", gas_virtual_rate / KBTU_PER_UNIT["THERM"])
    file_map[UTILITY_RATES_SHEET_XML_PATH] = ET.tostring(utility_root, encoding="utf-8", xml_declaration=True)


def populate_master_room_list_space_type_table(
    sim_text: str,
    workbook_path: Path,
    output_workbook_path: Path,
) -> Dict[str, object]:
    """Populate Master Room List 'Space Type Table' and Utilities table from LV-B + HOURLY + ES-D."""
    lv_b_result = extract_lv_b_spaces(sim_text)
    hourly_result = extract_hourly_thermostat_setpoint_ranges(sim_text)
    try:
        es_d_result = extract_es_d_energy_cost_summary(sim_text)
    except ValueError:
        es_d_result = {"utility_rates": {}}
    hourly_by_space = {
        _normalize_space_name(space_name): values
        for space_name, values in hourly_result["spaces"].items()
    }
    spaces = list(lv_b_result["spaces"].items())
    if not spaces:
        raise ValueError("No LV-B spaces found to populate the Master Room List.")
    if load_workbook is not None:
        workbook = load_workbook(workbook_path, keep_vba=True)
        sheet = workbook["Master Room List"]
        utility_sheet = workbook["Utilities"]
        max_rows = MASTER_ROOM_LIST_SPACE_MAX_ROWS
        start_row = MASTER_ROOM_LIST_SPACE_START_ROW
        for index in range(max_rows):
            row_number = start_row + index
            if index < len(spaces):
                space_name, space_data = spaces[index]
                thermostat_data = hourly_by_space.get(_normalize_space_name(space_name), {})
                sheet[f"D{row_number}"] = space_name
                sheet[f"G{row_number}"] = float(space_data["area_sqft"])
                sheet[f"H{row_number}"] = float(space_data["lights_w_per_sqft"])
                sheet[f"I{row_number}"] = float(space_data["equip_w_per_sqft"])
                sheet[f"J{row_number}"] = float(space_data["people"])
                sheet[f"K{row_number}"] = thermostat_data.get("max_thermostat_setpoint_f")
                sheet[f"L{row_number}"] = thermostat_data.get("min_thermostat_setpoint_f")
            else:
                sheet[f"D{row_number}"] = None
                sheet[f"G{row_number}"] = None
                sheet[f"H{row_number}"] = None
                sheet[f"I{row_number}"] = None
                sheet[f"J{row_number}"] = None
                sheet[f"K{row_number}"] = None
                sheet[f"L{row_number}"] = None
        utility_rates = es_d_result["utility_rates"]
        elec_virtual_rate = next(
            float(data["virtual_rate"]) for data in utility_rates.values() if str(data["unit"]).upper() == "KWH"
        )
        gas_virtual_rate = next(
            float(data["virtual_rate"]) for data in utility_rates.values() if str(data["unit"]).upper() == "THERM"
        )
        utility_sheet["B2"] = "Electrical"
        utility_sheet["C2"] = "kWh"
        utility_sheet["D2"] = elec_virtual_rate
        utility_sheet["E2"] = elec_virtual_rate / KBTU_PER_UNIT["KWH"]
        utility_sheet["B3"] = "Natural Gas"
        utility_sheet["C3"] = "Therms"
        utility_sheet["D3"] = gas_virtual_rate
        utility_sheet["E3"] = gas_virtual_rate / KBTU_PER_UNIT["THERM"]
        workbook.save(output_workbook_path)
        return {
            "target_sheet": "Master Room List",
            "target_table": "Space Type Table",
            "writer": "openpyxl",
            "rows_available": max_rows,
            "spaces_found": len(spaces),
            "spaces_written": min(len(spaces), max_rows),
            "spaces_truncated": max(len(spaces) - max_rows, 0),
            "utilities_updated": True,
            "output_workbook": str(output_workbook_path),
        }
    ET.register_namespace("", MAIN_NS)
    file_map = _load_zip_file_map(workbook_path)
    if MASTER_ROOM_LIST_SHEET_XML_PATH not in file_map:
        raise ValueError("Could not find Master Room List worksheet XML in the workbook.")
    sheet_root = _parse_xml_with_registered_namespaces(file_map[MASTER_ROOM_LIST_SHEET_XML_PATH])
    sheet_data = sheet_root.find("m:sheetData", NS)
    if sheet_data is None:
        raise ValueError("Workbook sheet is missing sheetData.")
    text_style_template = None
    area_style_template = None
    lpd_style_template = None
    epd_style_template = None
    occupancy_style_template = None
    setpoint_style_template = None
    setback_style_template = None
    text_template_cell = sheet_root.find(".//m:c[@r='E16']", NS)
    area_template_cell = sheet_root.find(".//m:c[@r='G16']", NS)
    lpd_template_cell = sheet_root.find(".//m:c[@r='H16']", NS)
    epd_template_cell = sheet_root.find(".//m:c[@r='I16']", NS)
    occupancy_template_cell = sheet_root.find(".//m:c[@r='J16']", NS)
    setpoint_template_cell = sheet_root.find(".//m:c[@r='K16']", NS)
    setback_template_cell = sheet_root.find(".//m:c[@r='L16']", NS)
    if text_template_cell is not None:
        text_style_template = text_template_cell.attrib.get("s")
    if area_template_cell is not None:
        area_style_template = area_template_cell.attrib.get("s")
    if lpd_template_cell is not None:
        lpd_style_template = lpd_template_cell.attrib.get("s")
    if epd_template_cell is not None:
        epd_style_template = epd_template_cell.attrib.get("s")
    if occupancy_template_cell is not None:
        occupancy_style_template = occupancy_template_cell.attrib.get("s")
    if setpoint_template_cell is not None:
        setpoint_style_template = setpoint_template_cell.attrib.get("s")
    if setback_template_cell is not None:
        setback_style_template = setback_template_cell.attrib.get("s")
    max_rows = MASTER_ROOM_LIST_SPACE_MAX_ROWS
    start_row = MASTER_ROOM_LIST_SPACE_START_ROW
    for index in range(max_rows):
        row_number = start_row + index
        row = _ensure_row(sheet_data, row_number)
        name_ref = f"D{row_number}"
        area_ref = f"G{row_number}"
        lpd_ref = f"H{row_number}"
        epd_ref = f"I{row_number}"
        occupants_ref = f"J{row_number}"
        setpoint_ref = f"K{row_number}"
        setback_ref = f"L{row_number}"
        if index < len(spaces):
            space_name, space_data = spaces[index]
            thermostat_data = hourly_by_space.get(_normalize_space_name(space_name), {})
            _set_inline_string_cell(row, name_ref, space_name, style=text_style_template)
            _set_numeric_cell(row, area_ref, float(space_data["area_sqft"]), style=area_style_template)
            _set_numeric_cell(row, lpd_ref, float(space_data["lights_w_per_sqft"]), style=lpd_style_template)
            _set_numeric_cell(row, epd_ref, float(space_data["equip_w_per_sqft"]), style=epd_style_template)
            _set_numeric_cell(row, occupants_ref, float(space_data["people"]), style=occupancy_style_template)
            _set_numeric_cell(
                row,
                setpoint_ref,
                thermostat_data.get("max_thermostat_setpoint_f"),
                style=setpoint_style_template,
            )
            _set_numeric_cell(
                row,
                setback_ref,
                thermostat_data.get("min_thermostat_setpoint_f"),
                style=setback_style_template,
            )
        else:
            _set_inline_string_cell(row, name_ref, "", style=text_style_template)
            _set_numeric_cell(row, area_ref, None, style=area_style_template)
            _set_numeric_cell(row, lpd_ref, None, style=lpd_style_template)
            _set_numeric_cell(row, epd_ref, None, style=epd_style_template)
            _set_numeric_cell(row, occupants_ref, None, style=occupancy_style_template)
            _set_numeric_cell(row, setpoint_ref, None, style=setpoint_style_template)
            _set_numeric_cell(row, setback_ref, None, style=setback_style_template)
    file_map[MASTER_ROOM_LIST_SHEET_XML_PATH] = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
    _write_utility_rate_table_from_es_d(file_map, es_d_result)
    _save_zip_file_map(file_map, output_workbook_path)
    return {
        "target_sheet": "Master Room List",
        "target_table": "Space Type Table",
        "rows_available": max_rows,
        "spaces_found": len(spaces),
        "spaces_written": min(len(spaces), max_rows),
        "spaces_truncated": max(len(spaces) - max_rows, 0),
        "utilities_updated": True,
        "output_workbook": str(output_workbook_path),
    }


def populate_ecm_data_from_reports(
    sim_text: str,
    workbook_path: Path,
    model_run_type: str,
    output_workbook_path: Path,
) -> Dict[str, object]:
    """Populate ECM Data sheet from BEPS + ES-D results for Baseline/Proposed/ECM-1..7."""
    normalized_model_run_type = model_run_type.strip().upper()
    if normalized_model_run_type not in ECM_DATA_MODEL_START_ROWS:
        raise ValueError(
            "Unsupported model run type for ECM Data. Supported: Baseline, Proposed, ECM-1..ECM-7 "
            "(Baseline-2 and Baseline-3 are intentionally ignored)."
        )
    beps_result = extract_beps_report(sim_text)
    try:
        es_d_result = extract_es_d_energy_cost_summary(sim_text)
    except ValueError:
        es_d_result = {"utility_rates": {}}
    elec_totals = beps_result["totals_by_fuel"]["electricity"]
    gas_totals = beps_result["totals_by_fuel"]["natural_gas"]
    elec_unit = elec_totals["unit"]
    gas_unit = gas_totals["unit"]
    elec_end_use_values_kbtu: Dict[str, float] = {}
    gas_end_use_values_kbtu: Dict[str, float] = {}
    for beps_column, target_column in BEPS_TO_ECM_END_USE_COLUMNS.items():
        elec_value = float(elec_totals["by_end_use"][beps_column])
        gas_value = float(gas_totals["by_end_use"][beps_column])
        elec_end_use_values_kbtu[target_column] = _to_kbtu(elec_value, elec_unit) if abs(elec_value) > 1e-9 else 0.0
        gas_end_use_values_kbtu[target_column] = _to_kbtu(gas_value, gas_unit) if abs(gas_value) > 1e-9 else 0.0
    total_electric_kbtu = _to_kbtu(float(elec_totals["by_end_use"]["TOTAL"]), elec_unit)
    total_gas_kbtu = _to_kbtu(float(gas_totals["by_end_use"]["TOTAL"]), gas_unit)
    total_energy_kbtu = total_electric_kbtu + total_gas_kbtu
    utility_rates = es_d_result["utility_rates"]
    elec_virtual_rate = None
    gas_virtual_rate = None
    for rate_data in utility_rates.values():
        unit = str(rate_data["unit"]).upper()
        if unit == "KWH" and elec_virtual_rate is None:
            elec_virtual_rate = float(rate_data["virtual_rate"])
        elif unit == "THERM" and gas_virtual_rate is None:
            gas_virtual_rate = float(rate_data["virtual_rate"])
    if elec_virtual_rate is None:
        elec_virtual_rate = 0.0
    if gas_virtual_rate is None:
        gas_virtual_rate = 0.0
    elec_rate_per_kbtu = elec_virtual_rate / KBTU_PER_UNIT["KWH"]
    gas_rate_per_kbtu = gas_virtual_rate / KBTU_PER_UNIT["THERM"]
    elec_cost = total_electric_kbtu * elec_rate_per_kbtu
    gas_cost = total_gas_kbtu * gas_rate_per_kbtu
    total_cost = elec_cost + gas_cost
    section_start = ECM_DATA_MODEL_START_ROWS[normalized_model_run_type]
    elec_energy_row_number = section_start + 1
    elec_demand_row_number = section_start + 2
    gas_energy_row_number = section_start + 3
    gas_demand_row_number = section_start + 4
    writable_columns = "BCDEFGHIJKLMNOPQRS"
    if load_workbook is not None:
        workbook = load_workbook(workbook_path, keep_vba=True)
        sheet = workbook["ECM Data"]
        for col in writable_columns:
            sheet[f"{col}{elec_energy_row_number}"] = None
            sheet[f"{col}{elec_demand_row_number}"] = None
            sheet[f"{col}{gas_energy_row_number}"] = None
            sheet[f"{col}{gas_demand_row_number}"] = None
        for col, value in elec_end_use_values_kbtu.items():
            sheet[f"{col}{elec_energy_row_number}"] = value
        for col, value in gas_end_use_values_kbtu.items():
            sheet[f"{col}{gas_energy_row_number}"] = value
        workbook.save(output_workbook_path)
        return {
            "sheet": "ECM Data",
            "writer": "openpyxl",
            "target_table": f"ECMData_{normalized_model_run_type.replace('-', '')}" if normalized_model_run_type.startswith("ECM-") else f"ECMData_{normalized_model_run_type.title()}",
            "model_run_type": model_run_type,
            "section_start_row": section_start,
            "electrical_energy_row": elec_energy_row_number,
            "natural_gas_energy_row": gas_energy_row_number,
            "end_use_columns_written": sorted(elec_end_use_values_kbtu.keys()),
            "left_blank_columns": ECM_OPTIONAL_BLANK_COLUMNS,
            "total_electric_kbtu": total_electric_kbtu,
            "total_gas_kbtu": total_gas_kbtu,
            "total_energy_kbtu": total_energy_kbtu,
            "elec_virtual_rate_per_kbtu": elec_rate_per_kbtu,
            "gas_virtual_rate_per_kbtu": gas_rate_per_kbtu,
            "elec_cost": elec_cost,
            "gas_cost": gas_cost,
            "total_cost": total_cost,
            "output_workbook": str(output_workbook_path),
        }
    ET.register_namespace("", MAIN_NS)
    file_map = _load_zip_file_map(workbook_path)
    if ECM_DATA_SHEET_XML_PATH not in file_map:
        raise ValueError("Could not find ECM Data worksheet XML in workbook.")
    sheet_root = _parse_xml_with_registered_namespaces(file_map[ECM_DATA_SHEET_XML_PATH])
    sheet_data = sheet_root.find("m:sheetData", NS)
    if sheet_data is None:
        raise ValueError("ECM Data sheet is missing sheetData.")
    elec_energy_row = _ensure_row(sheet_data, elec_energy_row_number)
    elec_demand_row = _ensure_row(sheet_data, elec_demand_row_number)
    gas_energy_row = _ensure_row(sheet_data, gas_energy_row_number)
    gas_demand_row = _ensure_row(sheet_data, gas_demand_row_number)
    for col in writable_columns:
        _set_numeric_cell(elec_energy_row, f"{col}{elec_energy_row_number}", None)
        _set_numeric_cell(elec_demand_row, f"{col}{elec_demand_row_number}", None)
        _set_numeric_cell(gas_energy_row, f"{col}{gas_energy_row_number}", None)
        _set_numeric_cell(gas_demand_row, f"{col}{gas_demand_row_number}", None)
    for col, value in elec_end_use_values_kbtu.items():
        _set_numeric_cell(elec_energy_row, f"{col}{elec_energy_row_number}", value)
    for col, value in gas_end_use_values_kbtu.items():
        _set_numeric_cell(gas_energy_row, f"{col}{gas_energy_row_number}", value)
    file_map[ECM_DATA_SHEET_XML_PATH] = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
    _save_zip_file_map(file_map, output_workbook_path)
    return {
        "sheet": "ECM Data",
        "target_table": f"ECMData_{normalized_model_run_type.replace('-', '')}" if normalized_model_run_type.startswith("ECM-") else f"ECMData_{normalized_model_run_type.title()}",
        "model_run_type": model_run_type,
        "section_start_row": section_start,
        "electrical_energy_row": elec_energy_row_number,
        "natural_gas_energy_row": gas_energy_row_number,
        "end_use_columns_written": sorted(elec_end_use_values_kbtu.keys()),
        "left_blank_columns": ECM_OPTIONAL_BLANK_COLUMNS,
        "total_electric_kbtu": total_electric_kbtu,
        "total_gas_kbtu": total_gas_kbtu,
        "total_energy_kbtu": total_energy_kbtu,
        "elec_virtual_rate_per_kbtu": elec_rate_per_kbtu,
        "gas_virtual_rate_per_kbtu": gas_rate_per_kbtu,
        "elec_cost": elec_cost,
        "gas_cost": gas_cost,
        "total_cost": total_cost,
        "output_workbook": str(output_workbook_path),
    }


def check_master_room_list_space_type_table_match(sim_text: str, workbook_path: Path) -> Dict[str, object]:
    """Compare LV-B spaces against existing Master Room List Space Type Table values."""
    lv_b_result = extract_lv_b_spaces(sim_text)
    expected_spaces = list(lv_b_result["spaces"].items())[:MASTER_ROOM_LIST_SPACE_MAX_ROWS]
    if load_workbook is not None:
        workbook = load_workbook(workbook_path, keep_vba=True, data_only=True)
        sheet = workbook["Master Room List"]
        mismatches: List[Dict[str, object]] = []
        compared_rows = 0
        for index in range(MASTER_ROOM_LIST_SPACE_MAX_ROWS):
            row_number = MASTER_ROOM_LIST_SPACE_START_ROW + index
            existing_name = sheet[f"D{row_number}"].value
            existing_area = sheet[f"G{row_number}"].value
            existing_name = str(existing_name).strip() if existing_name is not None else ""
            existing_area = float(existing_area) if existing_area is not None else None
            if index < len(expected_spaces):
                expected_name, expected_space_data = expected_spaces[index]
                expected_area = float(expected_space_data["area_sqft"])
                compared_rows += 1
                name_matches = existing_name == expected_name
                area_matches = existing_area is not None and abs(existing_area - expected_area) < 1e-6
                if not (name_matches and area_matches):
                    mismatches.append(
                        {
                            "row": row_number,
                            "expected_space_name": expected_name,
                            "existing_space_name": existing_name,
                            "expected_area_sqft": expected_area,
                            "existing_area_sqft": existing_area,
                        }
                    )
        return {
            "target_sheet": "Master Room List",
            "target_table": "Space Type Table",
            "writer": "openpyxl",
            "rows_checked": compared_rows,
            "spaces_compared": len(expected_spaces),
            "space_type_table_match": len(mismatches) == 0,
            "mismatch_count": len(mismatches),
            "mismatches": mismatches,
        }
    sheet_root = _load_master_room_list_sheet(workbook_path)
    sheet_data = sheet_root.find("m:sheetData", NS)
    if sheet_data is None:
        raise ValueError("Workbook sheet is missing sheetData.")
    mismatches: List[Dict[str, object]] = []
    compared_rows = 0
    for index in range(MASTER_ROOM_LIST_SPACE_MAX_ROWS):
        row_number = MASTER_ROOM_LIST_SPACE_START_ROW + index
        row = sheet_data.find(f"m:row[@r='{row_number}']", NS)
        existing_name = ""
        existing_area = None
        if row is not None:
            existing_name = _read_cell_text(row, f"D{row_number}").strip()
            existing_area = _read_cell_float(row, f"G{row_number}")
        if index < len(expected_spaces):
            expected_name, expected_space_data = expected_spaces[index]
            expected_area = float(expected_space_data["area_sqft"])
            compared_rows += 1
            name_matches = existing_name == expected_name
            area_matches = existing_area is not None and abs(existing_area - expected_area) < 1e-6
            if not (name_matches and area_matches):
                mismatches.append(
                    {
                        "row": row_number,
                        "expected_space_name": expected_name,
                        "existing_space_name": existing_name,
                        "expected_area_sqft": expected_area,
                        "existing_area_sqft": existing_area,
                    }
                )
    return {
        "target_sheet": "Master Room List",
        "target_table": "Space Type Table",
        "rows_checked": compared_rows,
        "spaces_compared": len(expected_spaces),
        "space_type_table_match": len(mismatches) == 0,
        "mismatch_count": len(mismatches),
        "mismatches": mismatches,
    }


def resolve_model_run_type(cli_model_run_type: str | None) -> str:
    """Resolve model run type from CLI first, then environment, defaulting to Baseline."""
    if cli_model_run_type and cli_model_run_type.strip():
        return cli_model_run_type.strip()
    env_value = os.getenv("MODEL_RUN_TYPE", "").strip()
    if env_value:
        return env_value
    return "Baseline"
def extract_es_d_energy_cost_summary(sim_text: str) -> Dict[str, object]:
    """Extract utility-rate virtual rate, metered unit, and total charge from ES-D."""
    lines = sim_text.splitlines()
    in_esd = False
    utility_rates: Dict[str, Dict[str, object]] = {}
    for raw_line in lines:
        stripped = raw_line.strip()
        upper = stripped.upper()
        if "REPORT- ES-D" in upper:
            in_esd = True
            continue
        if in_esd and upper.startswith("REPORT-") and "REPORT- ES-D" not in upper:
            in_esd = False
        if not in_esd:
            continue
        if (
            not stripped
            or upper.startswith("UTILITY-RATE")
            or upper.startswith("METERED")
            or upper.startswith("ENERGY COST/")
            or set(stripped) <= {"-", "=", "+"}
        ):
            continue
        parts = [part.strip() for part in re.split(r"\s{2,}", stripped) if part.strip()]
        if len(parts) < 6:
            continue
        utility_rate = parts[0]
        metered_energy = parts[-4]
        total_charge_str = parts[-3]
        virtual_rate_str = parts[-2]
        metered_tokens = metered_energy.split()
        if len(metered_tokens) < 2:
            continue
        unit = metered_tokens[-1]
        try:
            total_charge = float(total_charge_str.replace(",", ""))
            virtual_rate = float(virtual_rate_str.replace(",", ""))
        except ValueError:
            continue
        utility_rates[utility_rate] = {
            "unit": unit,
            "total_charge": total_charge,
            "total_charge_unit": "$",
            "virtual_rate": virtual_rate,
            "virtual_rate_unit": "$/Unit",
        }
    if not utility_rates:
        raise ValueError("Could not parse ES-D utility-rate rows from the SIM file.")
    return {
        "report": "ES-D",
        "utility_rates": utility_rates,
    }
def extract_ps_h_details(sim_text: str) -> Dict[str, object]:
    """Extract PS-H loop, pump, and equipment sizing details."""
    lines = sim_text.splitlines()
    loops: Dict[str, Dict[str, object]] = {}
    pumps: Dict[str, Dict[str, object]] = {}
    equipment: Dict[str, Dict[str, object]] = {}
    report_indices = [i for i, line in enumerate(lines) if "REPORT- PS-H" in line.upper()]
    for start in report_indices:
        line = lines[start]
        name_match = re.search(r"REPORT-\s*PS-H\s+Loads and Energy Usage for\s+(.+?)\s+WEATHER FILE", line, re.IGNORECASE)
        if not name_match:
            continue
        report_name = " ".join(name_match.group(1).split())
        end = len(lines)
        for j in range(start + 1, len(lines)):
            if lines[j].strip().upper().startswith("REPORT-"):
                end = j
                break
        block = lines[start:end]
        block_text = "\n".join(block)
        if "DETAILED SIZING INFORMATION" in block_text:
            units_line_idx = next((i for i, b in enumerate(block) if "(MBTU/HR)" in b and "(HOURS)" in b and "(BTU/BTU)" in b), None)
            if units_line_idx is not None:
                units = re.findall(r"\(([^\)]+)\)", block[units_line_idx])
                row = None
                for k in range(units_line_idx + 1, min(units_line_idx + 20, len(block))):
                    if "----" in block[k]:
                        continue
                    parts = [part.strip() for part in re.split(r"\s{2,}", block[k].strip()) if part.strip()]
                    if len(parts) >= 8 and re.fullmatch(r"-?\d+(?:\.\d+)?", parts[1]):
                        row = parts
                        break
                if row:
                    equipment[report_name] = {
                        "capacity": float(row[1]),
                        "start_up": float(row[2]),
                        "electric": float(row[3]),
                        "heat_eir": float(row[4]),
                        "aux_elec": float(row[5]),
                        "fuel": float(row[6]),
                        "heat_hir": float(row[7]),
                        "units": {
                            "capacity": units[0] if len(units) > 0 else "MBTU/HR",
                            "start_up": units[1] if len(units) > 1 else "HOURS",
                            "electric": units[2] if len(units) > 2 else "KW",
                            "heat_eir": units[3] if len(units) > 3 else "BTU/BTU",
                            "aux_elec": units[4] if len(units) > 4 else "KW",
                            "fuel": units[5] if len(units) > 5 else "MBTU/HR",
                            "heat_hir": units[6] if len(units) > 6 else "BTU/BTU",
                        },
                    }
        elif "HEATING     COOLING      LOOP" in block_text:
            # first PS-H loop instance table at top
            units_line_idx = next((i for i, b in enumerate(block) if "(MBTU/HR)" in b and "(GPM" in b and "(FT)" in b), None)
            if units_line_idx is None:
                continue
            units = re.findall(r"\(([^\)]+)\)", block[units_line_idx])
            value_parts = None
            for k in range(units_line_idx + 1, min(units_line_idx + 12, len(block))):
                candidate = block[k].strip()
                if not candidate or '----' in candidate:
                    continue
                parts = candidate.split()
                if len(parts) >= 10 and all(re.fullmatch(r"-?\d+(?:\.\d+)?", p) for p in parts[:10]):
                    value_parts = parts[:10]
                    break
            if value_parts:
                loops[report_name] = {
                    "heating_capacity": float(value_parts[0]),
                    "cooling_capacity": float(value_parts[1]),
                    "loop_flow": float(value_parts[2]),
                    "total_head": float(value_parts[3]),
                    "loop_volume": float(value_parts[8]),
                    "units": {
                        "heating_capacity": units[0] if len(units) > 0 else "MBTU/HR",
                        "cooling_capacity": units[1] if len(units) > 1 else "MBTU/HR",
                        "loop_flow": units[2] if len(units) > 2 else "GPM",
                        "total_head": units[3] if len(units) > 3 else "FT",
                        "loop_volume": units[8] if len(units) > 8 else "GAL",
                    },
                }
        elif "CAPACITY               MECHANICAL" in block_text and "ATTACHED TO" in block_text:
            header_idx = next((i for i, b in enumerate(block) if "ATTACHED TO" in b and "(GPM" in b and "(KW)" in b), None)
            if header_idx is None:
                continue
            units_line = block[header_idx]
            units = re.findall(r"\(([^\)]+)\)", units_line)
            row = None
            for k in range(header_idx + 1, min(header_idx + 12, len(block))):
                candidate = block[k].strip()
                if not candidate or '----' in candidate:
                    continue
                parts = [part.strip() for part in re.split(r"\s{2,}", candidate) if part.strip()]
                if len(parts) >= 8 and re.fullmatch(r"-?\d+(?:\.\d+)?", parts[1]):
                    row = parts
                    break
            if row:
                pumps[report_name] = {
                    "attached_to": row[0],
                    "flow": float(row[1]),
                    "head": float(row[2]),
                    "capacity_control": row[4],
                    "power": float(row[5]),
                    "mechanical_efficiency": float(row[6]),
                    "motor_efficiency": float(row[7]),
                    "units": {
                        "flow": units[0] if len(units) > 0 else "GPM",
                        "head": units[1] if len(units) > 1 else "FT",
                        "capacity_control": "unitless",
                        "power": units[3] if len(units) > 3 else "KW",
                        "mechanical_efficiency": units[4] if len(units) > 4 else "FRAC",
                        "motor_efficiency": units[5] if len(units) > 5 else "FRAC",
                    },
                }
    if not (loops or pumps or equipment):
        raise ValueError("Could not parse PS-H loop/pump/equipment details from the SIM file.")
    return {
        "report": "PS-H",
        "loops": loops,
        "pumps": pumps,
        "equipment": equipment,
    }


def extract_schedule_table(sim_text: str) -> Dict[str, object]:
    """Extract REPORT- SCHEDULES rows into a table-like structure."""
    lines = sim_text.splitlines()
    in_schedules = False
    header: List[str] | None = None
    rows: List[Dict[str, str]] = []
    for raw_line in lines:
        stripped = raw_line.strip()
        upper = stripped.upper()
        if "REPORT- SCHEDULES" in upper:
            in_schedules = True
            header = None
            continue
        if in_schedules and upper.startswith("REPORT-") and "REPORT- SCHEDULES" not in upper:
            break
        if not in_schedules or not stripped:
            continue
        if header is None and "SCHEDULE NAME" in upper and "SCHEDULE TYPE" in upper:
            header = [part.strip() for part in re.split(r"\s{2,}", stripped) if part.strip()]
            continue
        if header is None:
            continue
        parts = [part.strip() for part in re.split(r"\s{2,}", stripped) if part.strip()]
        if len(parts) < len(header):
            continue
        row = {header[idx]: parts[idx] for idx in range(len(header))}
        rows.append(row)
    return {"report": "SCHEDULE", "rows": rows}


def populate_equest_schedule_importer_table(
    sim_text: str,
    workbook_path: Path,
    output_workbook_path: Path,
) -> Dict[str, object]:
    """Populate the eQuest Schedule Importer tab from REPORT- SCHEDULES."""
    schedule_result = extract_schedule_table(sim_text)
    rows = schedule_result["rows"]
    ET.register_namespace("", MAIN_NS)
    file_map = _load_zip_file_map(workbook_path)
    schedule_sheet_path = "xl/worksheets/sheet16.xml"
    if schedule_sheet_path not in file_map:
        raise ValueError("Could not find eQuest Schedule Importer worksheet XML in workbook.")
    sheet_root = _parse_xml_with_registered_namespaces(file_map[schedule_sheet_path])
    sheet_data = sheet_root.find("m:sheetData", NS)
    if sheet_data is None:
        raise ValueError("eQuest Schedule Importer sheet is missing sheetData.")
    start_row = 2
    max_rows = 79
    for idx in range(max_rows):
        row_number = start_row + idx
        row = _ensure_row(sheet_data, row_number)
        row_data = rows[idx] if idx < len(rows) else {}
        _set_inline_string_cell(row, f"A{row_number}", row_data.get("Schedule Name", ""))
        _set_inline_string_cell(row, f"B{row_number}", row_data.get("Schedule Type", ""))
        for hour in range(1, 25):
            col = _excel_column_name(13 + hour)
            value = row_data.get(str(hour))
            numeric_value = float(value) if value not in (None, "") else None
            _set_numeric_cell(row, f"{col}{row_number}", numeric_value)
    file_map[schedule_sheet_path] = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
    _save_zip_file_map(file_map, output_workbook_path)
    return {
        "target_sheet": "eQuest Schedule Importer",
        "target_table": "eQuest_Schedule_Importer",
        "rows_written": min(len(rows), max_rows),
        "rows_available": max_rows,
        "output_workbook": str(output_workbook_path),
    }
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract BEPS, LV-B, LV-D, LV-I, LS-A, LV-M, ES-D, and PS-H report data from an eQuest SIM file."
    )
    parser.add_argument("sim_file", type=Path, help="Path to the eQuest .SIM file")
    parser.add_argument(
        "--report",
        choices=["beps", "lv-b", "lv-d", "lv-i", "ls-a", "lv-m", "es-d", "ps-h", "all"],
        default="all",
        help="Which report(s) to extract (default: all)",
    )
    parser.add_argument(
        "--indent",
        type=int,
        default=2,
        help="JSON indentation level for output (default: 2)",
    )
    parser.add_argument(
        "--populate-master-room-list",
        type=Path,
        help="Path to Building Performance Assumptions .xlsm used for Space Type Table actions.",
    )
    parser.add_argument(
        "--output-workbook",
        type=Path,
        help="Output path for updated workbook (required for Baseline writes).",
    )
    parser.add_argument(
        "--model-run-type",
        type=str,
        help="Model run type from automation input (e.g., Baseline, Proposed, ECM-1).",
    )
    parser.add_argument(
        "--update-ecm-data",
        type=Path,
        help="Path to workbook .xlsm where ECM Data should be populated from BEPS/ES-D.",
    )
    parser.add_argument(
        "--populate-schedules",
        type=Path,
        help="Path to workbook .xlsm where eQuest Schedule Importer should be populated.",
    )
    parser.add_argument(
        "--list-reports",
        action="store_true",
        help="List discovered REPORT-* sections in the SIM file and exit.",
    )
    args = parser.parse_args()
    sim_text = args.sim_file.read_text(errors="ignore")
    if args.list_reports:
        print(json.dumps(detect_available_reports(sim_text), indent=args.indent))
        return
    if args.update_ecm_data:
        if not args.output_workbook:
            raise ValueError("--output-workbook is required when using --update-ecm-data.")
        model_run_type = resolve_model_run_type(args.model_run_type)
        result = populate_ecm_data_from_reports(
            sim_text=sim_text,
            workbook_path=args.update_ecm_data,
            model_run_type=model_run_type,
            output_workbook_path=args.output_workbook,
        )
        print(json.dumps(result, indent=args.indent))
        return
    if args.populate_schedules:
        if not args.output_workbook:
            raise ValueError("--output-workbook is required when using --populate-schedules.")
        result = populate_equest_schedule_importer_table(
            sim_text=sim_text,
            workbook_path=args.populate_schedules,
            output_workbook_path=args.output_workbook,
        )
        print(json.dumps(result, indent=args.indent))
        return
    if args.populate_master_room_list:
        model_run_type = resolve_model_run_type(args.model_run_type)
        normalized_model_run_type = model_run_type.upper()
        is_baseline = normalized_model_run_type == "BASELINE"
        if is_baseline:
            if not args.output_workbook:
                raise ValueError("--output-workbook is required for Baseline when using --populate-master-room-list.")
            result = populate_master_room_list_space_type_table(
                sim_text=sim_text,
                workbook_path=args.populate_master_room_list,
                output_workbook_path=args.output_workbook,
            )
            result.update(
                {
                    "model_run_type": model_run_type,
                    "is_baseline": True,
                    "space_type_table_match": True,
                }
            )
        else:
            comparison_result = check_master_room_list_space_type_table_match(
                sim_text=sim_text,
                workbook_path=args.populate_master_room_list,
            )
            result = {
                "model_run_type": model_run_type,
                "is_baseline": False,
                **comparison_result,
            }
        print(json.dumps(result, indent=args.indent))
        return
    if args.report == "beps":
        result = extract_beps_report(sim_text)
    elif args.report == "lv-b":
        result = extract_lv_b_spaces(sim_text)
    elif args.report == "lv-d":
        result = extract_lv_d_report(sim_text)
    elif args.report == "lv-i":
        result = extract_lv_i_constructions(sim_text)
    elif args.report == "ls-a":
        result = extract_ls_a_peak_loads(sim_text)
    elif args.report == "lv-m":
        result = extract_lv_m_conversions(sim_text)
    elif args.report == "es-d":
        result = extract_es_d_energy_cost_summary(sim_text)
    elif args.report == "ps-h":
        result = extract_ps_h_details(sim_text)
    else:
        lv_b_result = extract_lv_b_spaces(sim_text)
        lv_m_result = extract_lv_m_conversions(sim_text)
        result = {
            "beps": extract_beps_report(sim_text),
            "lv_b_spaces": lv_b_result,
            "lv_d": extract_lv_d_report(sim_text),
            "lv_i": extract_lv_i_constructions(sim_text),
            "ls_a_peak_loads": extract_ls_a_peak_loads(sim_text, lv_b_result=lv_b_result),
            "lv_m": lv_m_result,
            "es_d": extract_es_d_energy_cost_summary(sim_text),
            "ps_h": extract_ps_h_details(sim_text),
        }
    print(json.dumps(result, indent=args.indent))
if __name__ == "__main__":
    main()
